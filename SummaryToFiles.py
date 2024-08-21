import openpyxl
import tkinter as tk
from tkinter import filedialog


def find_first_empty_cell_in_column(ws, column, start_row):
    """Find the first empty cell in a specific column starting from a given row"""
    for row in range(start_row, ws.max_row):
        if ws.cell(row=row, column=column).value is None:
            return row
    return ws.max_row + 1


def update_files_with_column_data(main_file_path_def, target_dir_def):
    try:
        # Open the main workbook
        main_workbook = openpyxl.load_workbook(main_file_path_def)
        main_worksheet = main_workbook.active  # Assuming data starts from the first sheet

        # Get the number of rows with data in the main worksheet
        row_count = main_worksheet.max_row

        for i in range(2, row_count + 1):
            # Read the filename from column G (7) and data from columns C (3), D (4), and E (5)
            file_name = main_worksheet.cell(row=i, column=7).value
            data_c = main_worksheet.cell(row=i, column=3).value
            data_d = main_worksheet.cell(row=i, column=4).value
            data_e = main_worksheet.cell(row=i, column=5).value

            if file_name:
                target_file_path = f"{target_dir_def}/{file_name}.xlsx"

                try:
                    # Open the target workbook
                    target_workbook = openpyxl.load_workbook(target_file_path)
                    target_worksheet = target_workbook.active  # Assuming data is on the first sheet

                    # Find the first empty cell in column C after the 5th row in the target worksheet
                    target_row = find_first_empty_cell_in_column(target_worksheet, column=3, start_row=6)

                    # Write the data from C, D, and E into the new rows in the target worksheet
                    target_worksheet.cell(row=target_row, column=3, value=data_c)
                    target_worksheet.cell(row=target_row, column=4, value=data_d)
                    target_worksheet.cell(row=target_row, column=7, value=data_e)

                    # Save the target workbook
                    target_workbook.save(target_file_path)

                except FileNotFoundError:
                    print(f"Error: Unable to open the target workbook '{file_name}'.")

    except FileNotFoundError:
        print("Error: Unable to open the main workbook.")


def select_main_file():
    main_file_path_main = filedialog.askopenfilename(title="Select Main Excel File", filetypes=[("Excel files",
                                                                                                 "*.xlsx *.xls")])
    return main_file_path_main


def select_target_directory():
    target_dir_main = filedialog.askdirectory(title="Select Target Directory")
    return target_dir_main


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    main_file_path = select_main_file()
    if not main_file_path:
        print("No main file selected. Exiting.")
        exit()

    target_dir = select_target_directory()
    if not target_dir:
        print("No target directory selected. Exiting.")
        exit()

    update_files_with_column_data(main_file_path, target_dir)
